"""文件全文解析工具。提供给 KB 索引服务和 step1 共用。

返回纯文本：xlsx 序列化为「[sheet] col=val; ...」每行一条；其他类型读完整文本。
"""

from __future__ import annotations

import logging
from pathlib import Path

logger = logging.getLogger(__name__)

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover
    PdfReader = None  # type: ignore[assignment]

try:
    from docx import Document as DocxDocument
except Exception:  # pragma: no cover
    DocxDocument = None  # type: ignore[assignment]

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None  # type: ignore[assignment]


def parse_file_full(file_path: str | Path) -> tuple[str, str]:
    """读完整文本。返回 (text, file_type)，file_type 为去掉点的扩展名。"""

    path = Path(file_path)
    ext = path.suffix.lower().lstrip(".")

    try:
        if ext == "pdf":
            return _parse_pdf(path), ext
        if ext in {"docx", "doc"}:
            return _parse_docx(path), ext
        if ext in {"xlsx", "xls"}:
            return _parse_xlsx(path), ext
        if ext in {"md", "txt", "csv"}:
            return path.read_text(encoding="utf-8", errors="ignore"), ext
        # 未知类型：返回提示性文本，KB 调用方决定如何处理
        return f"暂不支持的文件类型: {path.name}", ext or "unknown"
    except Exception as exc:  # noqa: BLE001
        logger.warning("parse_file_full 失败 path=%s: %s", path, exc)
        return f"解析失败：{exc}", ext or "unknown"


def _parse_pdf(path: Path) -> str:
    if PdfReader is None:
        return ""
    reader = PdfReader(str(path))
    texts: list[str] = []
    for page in reader.pages:
        try:
            texts.append(page.extract_text() or "")
        except Exception:
            texts.append("")
    return "\n".join(t for t in texts if t).strip()


def _parse_docx(path: Path) -> str:
    if DocxDocument is None:
        return ""
    doc = DocxDocument(str(path))
    return "\n".join(p.text for p in doc.paragraphs if p.text and p.text.strip())


def _parse_xlsx(path: Path) -> str:
    if load_workbook is None:
        return ""
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    out: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(c).strip() if c is not None else "" for c in rows[0]]
        for row in rows[1:]:
            parts: list[str] = []
            for header, value in zip(headers, row, strict=False):
                if value is None:
                    continue
                parts.append(f"{header}={value}" if header else str(value))
            if parts:
                out.append(f"[{sheet_name}] " + "; ".join(parts))
    return "\n".join(out)
